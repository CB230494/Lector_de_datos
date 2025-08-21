# =========================
# 🛠️ Panel del Administrador – Asistencia (editar, eliminar, Excel oficial con plantilla)
# =========================
import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import date, time
import sqlite3
from pathlib import Path

st.set_page_config(page_title="Asistencia - Administrador", layout="wide")
st.markdown("# 🛠️ Panel del Administrador – Asistencia")

# ---------- DB (SQLite persistente) ----------
DB_PATH = "asistencia.db"

def get_conn():
    return sqlite3.connect(DB_PATH, check_same_thread=False, timeout=30)

def init_db():
    with get_conn() as conn:
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA synchronous=NORMAL;")
        conn.execute("""
        CREATE TABLE IF NOT EXISTS asistencia(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at TEXT NOT NULL DEFAULT (datetime('now','localtime')),
            nombre TEXT, cedula TEXT, institucion TEXT,
            cargo TEXT, telefono TEXT,
            genero TEXT, sexo TEXT, edad TEXT
        );
        """)

def fetch_all_df(include_id=True):
    with get_conn() as conn:
        df = pd.read_sql_query("""
            SELECT id,
                   nombre  AS 'Nombre',
                   cedula  AS 'Cédula de Identidad',
                   institucion AS 'Institución',
                   cargo   AS 'Cargo',
                   telefono AS 'Teléfono',
                   genero  AS 'Género',
                   sexo    AS 'Sexo',
                   edad    AS 'Rango de Edad'
            FROM asistencia
            ORDER BY id ASC
        """, conn)
    if not df.empty:
        df.insert(0, "Nº", range(1, len(df)+1))
    if not include_id and not df.empty:
        return df.drop(columns=["id"])
    return df

def insert_row(row):
    with get_conn() as conn:
        conn.execute("""INSERT INTO asistencia
            (nombre, cedula, institucion, cargo, telefono, genero, sexo, edad)
            VALUES (?,?,?,?,?,?,?,?)""",
            (row["Nombre"], row["Cédula de Identidad"], row["Institución"],
             row["Cargo"], row["Teléfono"], row["Género"], row["Sexo"], row["Rango de Edad"])
        )

def update_row_by_id(row_id:int, row:dict):
    with get_conn() as conn:
        conn.execute("""
            UPDATE asistencia
               SET nombre=?, cedula=?, institucion=?, cargo=?, telefono=?, genero=?, sexo=?, edad=?
             WHERE id=?""",
            (row["Nombre"], row["Cédula de Identidad"], row["Institución"],
             row["Cargo"], row["Teléfono"], row["Género"], row["Sexo"], row["Rango de Edad"], row_id)
        )

def delete_rows_by_ids(ids):
    if not ids:
        return
    with get_conn() as conn:
        q = ",".join("?" for _ in ids)
        conn.execute(f"DELETE FROM asistencia WHERE id IN ({q})", ids)

def delete_all_rows():
    with get_conn() as conn:
        conn.execute("DELETE FROM asistencia;")

init_db()

# ---------- Tabs ----------
tab_form, tab_tabla, tab_excel = st.tabs(["📝 Formulario de Encabezado", "👥 Registros y edición", "⬇️ Excel oficial"])

# =========================
# 📝 1) Formulario de Encabezado (para el Excel oficial)
# =========================
with tab_form:
    st.caption("Estos datos rellenan los campos superiores de la plantilla oficial (con logos).")
    col1, col2 = st.columns([1,1])
    with col1:
        fecha_evento = st.date_input("Fecha", value=date.today())
        lugar = st.text_input("Lugar", value="sesión virtual")
        estrategia = st.text_input("Estrategia o Programa", value="Estrategia Sembremos Seguridad")
    with col2:
        from datetime import time as _t
        hora_inicio = st.time_input("Hora Inicio", value=_t(9,0))
        hora_fin = st.time_input("Hora Finalización", value=_t(12,0))
        delegacion = st.text_input("Dirección / Delegación Policial", value="Naranjo")

    st.info("Cuando descargues el Excel, se aplicarán estos datos en la plantilla.")

# =========================
# 👥 2) Registros: ver, editar, eliminar
# =========================
with tab_tabla:
    st.markdown("### 👥 Registros recibidos")
    df_all = fetch_all_df(include_id=True)

    if df_all.empty:
        st.info("Aún no hay registros guardados.")
    else:
        editable = df_all.copy()
        editable["Seleccionar"] = False

        edited = st.data_editor(
            editable[["Nº","Nombre","Cédula de Identidad","Institución","Cargo","Teléfono",
                      "Género","Sexo","Rango de Edad","Seleccionar"]],
            hide_index=True,
            use_container_width=True,
            column_config={
                "Nº": st.column_config.NumberColumn("Nº", disabled=True),
                "Seleccionar": st.column_config.CheckboxColumn("Seleccionar", help="Marca para eliminar"),
                "Género": st.column_config.SelectboxColumn("Género", options=["F","M","LGBTIQ+"]),
                "Sexo": st.column_config.SelectboxColumn("Sexo", options=["H","M","I"]),
                "Rango de Edad": st.column_config.SelectboxColumn("Rango de Edad",
                    options=["18 a 35 años","36 a 64 años","65 años o más"])
            },
            key="tabla_admin_editable"
        )

        c1, c2, c3, c4 = st.columns([1.2, 1.2, 1.2, 2])
        btn_save = c1.button("💾 Guardar cambios", use_container_width=True)
        btn_delete = c2.button("🗑️ Eliminar seleccionados", use_container_width=True)
        confirm_all = c4.checkbox("Confirmar vaciado total", value=False)
        btn_clear = c3.button("🧹 Vaciar todos", use_container_width=True)

        if btn_save:
            changes = 0
            for idx in edited.index:
                if idx >= len(df_all):
                    continue
                row_orig = df_all.loc[idx]
                row_new = edited.loc[idx]
                fields = ["Nombre","Cédula de Identidad","Institución","Cargo","Teléfono","Género","Sexo","Rango de Edad"]
                diff = any(str(row_orig[f]) != str(row_new[f]) for f in fields)
                if diff:
                    update_row_by_id(
                        int(row_orig["id"]),
                        {f: row_new[f] for f in fields}
                    )
                    changes += 1
            if changes:
                st.success(f"Se guardaron {changes} cambio(s).")
                st.rerun()
            else:
                st.info("No hay cambios para guardar.")

        if btn_delete:
            idx_sel = edited.index[edited["Seleccionar"] == True].tolist()
            ids = df_all.iloc[idx_sel]["id"].tolist()
            if ids:
                delete_rows_by_ids(ids)
                st.success(f"Eliminadas {len(ids)} fila(s).")
                st.rerun()
            else:
                st.info("No hay filas seleccionadas para eliminar.")

        if btn_clear:
            if confirm_all:
                delete_all_rows()
                st.success("Se vaciaron todos los registros.")
                st.rerun()
            else:
                st.warning("Marca la casilla 'Confirmar vaciado total' para continuar.")

# =========================
# ⬇️ 3) Excel oficial (usa la PLANTILLA directamente)
# =========================
with tab_excel:
    st.markdown("### ⬇️ Descargar Excel oficial con logos (usa tu plantilla tal cual)")
    st.caption("Se usa **LA-2025-NARANJO.xlsx** directamente. Se rellenan encabezados y tabla en la hoja 'Minuta'; si hay más de 16 registros, se generan hojas 'Minuta 2', 'Minuta 3', etc. dentro del **mismo archivo**.")
    TEMPLATE_PATH = "LA-2025-NARANJO.xlsx"   # deja este archivo junto al .py

    df_all = fetch_all_df(include_id=True)

    def _fecha_es(fecha: date) -> str:
        meses = ["enero","febrero","marzo","abril","mayo","junio",
                 "julio","agosto","septiembre","octubre","noviembre","diciembre"]
        return f"{fecha.day} {meses[fecha.month-1]} {fecha.year}"

    def build_excel_from_template(template_path: str,
                                  fecha: date,
                                  lugar: str,
                                  hora_ini: time,
                                  hora_fin: time,
                                  estrategia: str,
                                  delegacion: str,
                                  rows_df: pd.DataFrame) -> bytes:
        """
        Rellena la propia plantilla (sin recrear libro nuevo):
          - B6: 'Fecha: ...'
          - E6: 'Lugar: ...'
          - J6: 'Hora Inicio: ...'
          - Q6: 'Hora Finalización: ...'
          - B7: 'Estrategia o Programa: ...'
          - E8: delegación
        Tabla (desde fila 11 aprox., detectada por merges C:E):
          - Nombre en C:E (merge por fila)
          - F Cédula | G Institución | H Cargo | I Teléfono
          - J/K/L Género (F/M/LGBTIQ+) marcados con 'X'
          - M/N/O Sexo (H/M/I) marcados con 'X'
          - P/Q/R Rango de Edad (18-35 / 36-64 / 65+)
        Si hay más de 16 registros, se copian hojas 'Minuta' adicionales dentro del MISMO workbook.
        """
        try:
            from openpyxl import load_workbook
        except Exception:
            st.error("Falta 'openpyxl' en requirements.txt")
            return b""

        p = Path(template_path)
        if not p.exists():
            st.error(f"No se encontró la plantilla: {template_path}")
            return b""

        wb = load_workbook(p, data_only=False, keep_vba=False)
        if "Minuta" not in wb.sheetnames:
            st.error("La plantilla no contiene una hoja llamada 'Minuta'.")
            return b""

        ws_base = wb["Minuta"]

        # Detectar filas de la tabla por merges C:E (mantiene formato/logo)
        slots = []
        for r in ws_base.merged_cells.ranges:
            if r.min_col == 3 and r.max_col == 5 and r.min_row == r.max_row:
                slots.append(r.min_row)
        slots = sorted(slots)
        if not slots:
            st.error("No se detectaron filas de tabla en la plantilla (C:E merge por fila).")
            return b""
        start_row = slots[0]
        per_page = len(slots)  # normalmente 16

        # Particionar registros en páginas
        total = len(rows_df)
        pages = max(1, (total + per_page - 1) // per_page) if total > 0 else 1

        # Crear copias de 'Minuta' si se necesitan más páginas
        sheet_names = ["Minuta"]
        for p_i in range(1, pages):
            ws_copy = wb.copy_worksheet(ws_base)  # dentro del MISMO libro => conserva logos/estilos
            ws_copy.title = f"Minuta {p_i+1}"
            sheet_names.append(ws_copy.title)

        # Función para rellenar una hoja con su porción
        def fill_sheet(ws, df_slice: pd.DataFrame):
            # Encabezados (se llenan siempre para que quede exacto)
            ws["B6"].value = f"Fecha: {_fecha_es(fecha)}"
            ws["E6"].value = f"Lugar:  {lugar}"
            ws["J6"].value = f"Hora Inicio: {hora_ini.strftime('%H:%M')}"
            ws["Q6"].value = f"Hora Finalización: {hora_fin.strftime('%H:%M')}"
            ws["B7"].value = f"Estrategia o Programa: {estrategia}"
            ws["E8"].value = delegacion

            # Columnas de la tabla
            COL_NOMBRE_L = "C"  # merge C:E
            COL_CED = "F"
            COL_INST = "G"
            COL_CARGO = "H"
            COL_TEL = "I"
            COL_GEN_F, COL_GEN_M, COL_GEN_L = "J", "K", "L"
            COL_SEX_H, COL_SEX_M, COL_SEX_I = "M", "N", "O"
            COL_ED_1, COL_ED_2, COL_ED_3 = "P", "Q", "R"

            # Llenar filas
            for i, (_, row) in enumerate(df_slice.iterrows()):
                r = start_row + i
                ws[f"{COL_NOMBRE_L}{r}"].value = str(row["Nombre"]) if pd.notna(row["Nombre"]) else ""
                ws[f"{COL_CED}{r}"].value = str(row["Cédula de Identidad"] or "")
                ws[f"{COL_INST}{r}"].value = str(row["Institución"] or "")
                ws[f"{COL_CARGO}{r}"].value = str(row["Cargo"] or "")
                ws[f"{COL_TEL}{r}"].value = str(row["Teléfono"] or "")

                # Limpiar checkboxes en esa fila (no tocamos la columna 'S' FIRMA)
                for c in [COL_GEN_F, COL_GEN_M, COL_GEN_L, COL_SEX_H, COL_SEX_M, COL_SEX_I, COL_ED_1, COL_ED_2, COL_ED_3]:
                    ws[f"{c}{r}"].value = ""

                g = (row["Género"] or "").strip()
                if g == "F": ws[f"{COL_GEN_F}{r}"].value = "X"
                elif g == "M": ws[f"{COL_GEN_M}{r}"].value = "X"
                elif g == "LGBTIQ+": ws[f"{COL_GEN_L}{r}"].value = "X"

                s = (row["Sexo"] or "").strip()
                if s == "H": ws[f"{COL_SEX_H}{r}"].value = "X"
                elif s == "M": ws[f"{COL_SEX_M}{r}"].value = "X"
                elif s == "I": ws[f"{COL_SEX_I}{r}"].value = "X"

                e = (row["Rango de Edad"] or "").strip()
                if e.startswith("18"): ws[f"{COL_ED_1}{r}"].value = "X"
                elif e.startswith("36"): ws[f"{COL_ED_2}{r}"].value = "X"
                elif e.startswith("65"): ws[f"{COL_ED_3}{r}"].value = "X"

            # Opcional: limpiar restos de filas no usadas de esta página (excepto FIRMA en S)
            # for i in range(len(df_slice), per_page):
            #     r = start_row + i
            #     for c in [COL_NOMBRE_L, COL_CED, COL_INST, COL_CARGO, COL_TEL,
            #               COL_GEN_F, COL_GEN_M, COL_GEN_L, COL_SEX_H, COL_SEX_M, COL_SEX_I,
            #               COL_ED_1, COL_ED_2, COL_ED_3]:
            #         ws[f"{c}{r}"].value = ""

        # Rellenar cada página en su hoja correspondiente
        for p_i in range(pages):
            start = p_i * per_page
            end = min(start + per_page, total)
            df_slice = rows_df.iloc[start:end].reset_index(drop=True) if total > 0 else rows_df.head(0)
            ws = wb[sheet_names[p_i]]
            fill_sheet(ws, df_slice)

        bio = BytesIO(); wb.save(bio); return bio.getvalue()

    # Botón de descarga (usa plantilla)
    if st.button("📥 Generar y descargar Excel oficial", use_container_width=True, type="primary"):
        datos = df_all[["Nombre","Cédula de Identidad","Institución","Cargo","Teléfono","Género","Sexo","Rango de Edad"]] if not df_all.empty else df_all
        xls_bytes = build_excel_from_template(
            TEMPLATE_PATH,
            fecha_evento,
            lugar,
            hora_inicio,
            hora_fin,
            estrategia,
            delegacion,
            datos
        )
        if xls_bytes:
            st.download_button(
                "⬇️ Descargar Excel (plantilla oficial con logos)",
                data=xls_bytes,
                file_name=f"Lista_Asistencia_Oficial_{date.today():%Y%m%d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )




