# =========================
# 📋 Asistencia – Público + Admin (admin oculto hasta login)
# =========================
import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import date, time, datetime
from typing import List

st.set_page_config(page_title="Asistencia – Registro y Admin", layout="wide")

# ---------- Backend de datos: Google Sheets ----------
import gspread
try:
    from zoneinfo import ZoneInfo  # Python 3.9+
except Exception:
    ZoneInfo = None

# ID de tu Google Sheet (Asistencia Rivera)
SHEET_ID = "1lhREae4X-RcbeMmjSpT3CRJZo5enizyHmZxazzDGI-4"
SHEET_NAME = "Hoja 1"

# Estructura final (sin id/created_at)
HEADER = ["nombre","cedula","delegacion","cargo","telefono","genero","sexo","edad"]  # 8 columnas

# Catálogo de delegaciones (usado en el formulario público)
DELEGACIONES = [
    "Carmen","Merced","Hospital","Catedral","San Sebastián","Hatillo",
    "Zapote / San Francisco de dos Rios","Pavas","Uruca / Mata Redonda",
    "Curridabat","Montes de Oca","Goicoechea","Moravia","Tibás","Coronado",
    "Desamparados Norte","Desamparados Sur","Aserrí","Acosta","Alajuelita",
    "Escazu","Santa Ana","Mora","Puriscal","Turrubares",
    "Alajuela Sur","Alajuela Norte","San Ramón","Grecia","San Mateo",
    "Atenas","Naranjo","Palmares","Poas","Orotina","Sarchí",
    "Cartago","Paraíso","La Unión","Jiménez","Turrialba","Alvarado","Oreamuno","El Guarco",
    "Tarrazú","Dota","León Cortéz",
    "Guadalupe","Heredia","Barva","Santo Domingo","Santa Barbara","San Rafael","San Isidro","Belén","Flores","San Pablo",
    "Liberia","Nicoya","Santa Cruz","Bagaces","Carrillo","Cañas","Abangares","Tilarán","Nandayure","Hojancha","La Cruz",
    "Puntarenas","Esparza","Montes de Oro","Quepos","Parrita","Garabito","Paquera","Chomes",
    "Pérez Zeledón","Buenos Aires","Osa",
    "San Carlos Este","San Carlos Oeste","Zarcero","Guatuso","Río Cuarto",
    "Limón","Siquires","Talamanca","Matina",
    "Golfito","Coto Brus","Corredores","Puerto Jiménez",
    "Upala","Los Chiles - Cutris - Pocosol","Sarapiquí","Colorado","Pococí","Guacimo",
]

def _sa_key():
    try:
        sa = st.secrets["gcp_service_account"]
        return sa.get("client_email","") + "|" + sa.get("project_id","")
    except Exception:
        return ""

@st.cache_resource(show_spinner=False)
def _get_ws_cached(sheet_id: str, sheet_name: str, sa_key: str):
    if "gcp_service_account" not in st.secrets:
        raise RuntimeError("Falta el bloque [gcp_service_account] en .streamlit/secrets.toml")
    gc = gspread.service_account_from_dict(st.secrets["gcp_service_account"])
    sh = gc.open_by_key(sheet_id)

    # Obtiene/crea la pestaña
    try:
        ws = sh.worksheet(sheet_name)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=sheet_name, rows=2000, cols=len(HEADER))
        ws.update("A1:H1", [HEADER])
        try: ws.freeze(rows=1)
        except: pass

    # Migración automática: si aún existen columnas viejas (id/created_at), se eliminan A y B.
    try:
        first_row = [h.strip().lower() for h in ws.row_values(1)]
        if len(first_row) >= 2 and first_row[0] == "id" and first_row[1] == "created_at":
            ws.delete_columns(1, 2)  # borra A..B
    except Exception:
        pass

    # Asegura encabezado exacto
    first_row = [h.strip().lower() for h in ws.row_values(1)]
    if first_row != HEADER:
        ws.update("A1:H1", [HEADER])
        try: ws.freeze(rows=1)
        except: pass

    return ws

def _get_ws():
    return _get_ws_cached(SHEET_ID, SHEET_NAME, _sa_key())

def init_db():
    _get_ws()  # Garantiza existencia y encabezado/migración

def _now_local_str():
    try:
        tz = ZoneInfo("America/Costa_Rica")
        return datetime.now(tz).strftime("%Y-%m-%d %H:%M")
    except Exception:
        return datetime.now().strftime("%Y-%m-%d %H:%M")

# ---------- CRUD ----------
def insert_row(row: dict):
    ws = _get_ws()
    telefono = row.get("Teléfono","")
    if telefono and not str(telefono).startswith("'"):  # conserva ceros iniciales
        telefono = "'" + str(telefono)

    payload = [
        row.get("Nombre",""),
        row.get("Cédula de Identidad",""),
        row.get("Delegación",""),
        row.get("Cargo",""),
        telefono,
        row.get("Género",""),
        row.get("Sexo",""),
        row.get("Rango de Edad",""),
    ]
    ws.append_row(payload, value_input_option="USER_ENTERED")

def fetch_all_df(include_rownum=True) -> pd.DataFrame:
    """Lee todo y añade 'rownum' (número de fila real en el Sheet) para editar/borrar."""
    ws = _get_ws()
    values = ws.get_all_values()
    if len(values) < 2:
        cols = ["Nº","Nombre","Cédula de Identidad","Delegación","Cargo","Teléfono","Género","Sexo","Rango de Edad"]
        if include_rownum: cols.insert(1, "rownum")
        return pd.DataFrame(columns=cols)

    header = [h.strip().lower() for h in values[0]]
    data_rows = values[1:]

    name_map = {
        "nombre":"Nombre",
        "cedula":"Cédula de Identidad",
        "delegacion":"Delegación",
        "cargo":"Cargo",
        "telefono":"Teléfono",
        "genero":"Género",
        "sexo":"Sexo",
        "edad":"Rango de Edad",
    }

    records = []
    for idx, row in enumerate(data_rows, start=2):  # fila 2 es la primera de datos
        rec = {}
        for j, key in enumerate(header):
            if key in name_map:
                rec[name_map[key]] = row[j] if j < len(row) else ""
        rec["rownum"] = idx
        records.append(rec)

    df = pd.DataFrame(records)
    if df.empty:
        cols = ["Nº","Nombre","Cédula de Identidad","Delegación","Cargo","Teléfono","Género","Sexo","Rango de Edad"]
        if include_rownum: cols.insert(1, "rownum")
        return pd.DataFrame(columns=cols)

    cols_order = ["rownum","Nombre","Cédula de Identidad","Delegación","Cargo","Teléfono","Género","Sexo","Rango de Edad"]
    df = df[[c for c in cols_order if c in df.columns]]

    df.insert(0, "Nº", range(1, len(df)+1))
    if not include_rownum and "rownum" in df.columns:
        df = df.drop(columns=["rownum"])
    return df

def update_row_by_rownum(rownum:int, row:dict):
    ws = _get_ws()
    payload = [
        row.get("Nombre",""),
        row.get("Cédula de Identidad",""),
        row.get("Delegación",""),
        row.get("Cargo",""),
        row.get("Teléfono",""),
        row.get("Género",""),
        row.get("Sexo",""),
        row.get("Rango de Edad",""),
    ]
    ws.update(f"A{rownum}:H{rownum}", [payload], value_input_option="USER_ENTERED")

def delete_rows_by_rownums(rownums: List[int]):
    if not rownums:
        return
    ws = _get_ws()
    for r in sorted(rownums, reverse=True):
        ws.delete_rows(r)

def delete_all_rows():
    ws = _get_ws()
    used_rows = len(ws.get_all_values())
    if used_rows >= 2:
        ws.batch_clear([f"A2:H{used_rows}"])

# Inicializa backend (si falla, muestra error claro y detiene)
try:
    init_db()
except Exception:
    st.error("Error conectando a Google Sheets. Verifica permisos y secrets.")
    st.stop()

# ---------- Login admin en la barra lateral ----------
if "is_admin" not in st.session_state:
    st.session_state.is_admin = False

with st.sidebar:
    st.markdown("### 🔐 Acceso administrador")
    if not st.session_state.is_admin:
        pwd = st.text_input("Contraseña", type="password", placeholder="••••••••")
        if st.button("Ingresar"):
            if pwd == "Sembremos23":
                st.session_state.is_admin = True
                st.success("Acceso concedido.")
                st.rerun()
            else:
                st.error("Contraseña incorrecta.")
    else:
        st.success("Sesión de administrador activa")
        if st.button("Cerrar sesión"):
            st.session_state.is_admin = False
            st.rerun()

# ---------- Contenido PÚBLICO ----------
st.markdown("# 📋 Asistencia – Registro")
st.markdown("### ➕ Agregar")
with st.form("form_asistencia_publico", clear_on_submit=True):
    c1, c2, c3 = st.columns([1.2, 1, 1])
    nombre      = c1.text_input("Nombre")
    cedula      = c2.text_input("Cédula de Identidad")

    # Select en vez de texto libre para Delegación
    opciones_deleg = ["— Selecciona una delegación —"] + DELEGACIONES
    sel_deleg = c3.selectbox("Delegación", opciones_deleg, index=0)
    institucion = "" if sel_deleg == opciones_deleg[0] else sel_deleg

    c4, c5 = st.columns([1, 1])
    cargo    = c4.text_input("Cargo")
    telefono = c5.text_input("Teléfono")

    st.markdown("#### ")
    gcol, scol, ecol = st.columns([1.1, 1.5, 1.5])
    genero = gcol.radio("Género", ["F", "M", "LGBTIQ+"], horizontal=True)
    sexo   = scol.radio("Sexo (Hombre, Mujer o Intersex)", ["H", "M", "I"], horizontal=True)
    edad   = ecol.radio("Rango de Edad", ["18 a 35 años", "36 a 64 años", "65 años o más"], horizontal=True)

    submitted = st.form_submit_button("➕ Agregar", use_container_width=True)
    if submitted:
        if not nombre.strip():
            st.warning("Ingresa al menos el nombre.")
        else:
            fila = {
                "Nombre": nombre.strip(),
                "Cédula de Identidad": cedula.strip(),
                "Delegación": institucion.strip(),
                "Cargo": cargo.strip(),
                "Teléfono": telefono.strip(),
                "Género": genero,
                "Sexo": sexo,
                "Rango de Edad": edad
            }
            insert_row(fila)
            st.success("Registro guardado.")

st.markdown("### 📥 Registros recibidos")
df_pub = fetch_all_df(include_rownum=False)
if not df_pub.empty:
    st.dataframe(
        df_pub[["Nº","Nombre","Cédula de Identidad","Delegación","Cargo","Teléfono","Género","Sexo","Rango de Edad"]],
        use_container_width=True, hide_index=True
    )
else:
    st.info("Aún no hay registros guardados.")

# ---------- Contenido ADMIN ----------
if st.session_state.is_admin:
    st.markdown("---")
    st.markdown("# 🛠️ Panel del Administrador")

    # ==== Filtro por Delegación ====
    df_all = fetch_all_df(include_rownum=True)
    if df_all.empty:
        st.info("Aún no hay registros guardados.")
        st.stop()

    delegs_existentes = sorted([d for d in df_all["Delegación"].dropna().unique() if str(d).strip()], key=str.casefold)
    filtro_opts = ["(Todas)"] + delegs_existentes
    sel_filtro = st.selectbox("Filtrar por Delegación", filtro_opts, index=0)

    if sel_filtro == "(Todas)":
        df_view = df_all.copy().reset_index(drop=True)
    else:
        df_view = df_all[df_all["Delegación"] == sel_filtro].reset_index(drop=True)

    # Encabezado para Excel
    st.markdown("### 🧾 Datos de encabezado (Excel)")
    col1, col2 = st.columns([1,1])
    with col1:
        fecha_evento = st.date_input("Fecha", value=date.today())
        lugar = st.text_input("Lugar", value="")
        estrategia = st.text_input("Estrategia o Programa", value="Estrategia Sembremos Seguridad")
    with col2:
        hora_inicio = st.time_input("Hora Inicio", value=time(9,0))
        hora_fin = st.time_input("Hora Finalización", value=time(12,10))
        delegacion_hdr = st.text_input("Dirección / Delegación Policial", value=("" if sel_filtro == "(Todas)" else sel_filtro))
        firmante_nombre = st.text_input("Nombre de quien firma (opcional)", value="")

    st.markdown("### 📝 Anotaciones y Acuerdos (para el Excel)")
    a_col, b_col = st.columns(2)
    anotaciones = a_col.text_area("Anotaciones Generales", height=220, placeholder="Escribe las anotaciones generales…")
    acuerdos    = b_col.text_area("Acuerdos", height=220, placeholder="Escribe los acuerdos…")

    st.markdown("### 👥 Registros y edición")
    if df_view.empty:
        st.info("No hay registros para la delegación seleccionada.")
    else:
        editable = df_view.copy()
        editable["Seleccionar"] = False

        edited = st.data_editor(
            editable[["Nº","Nombre","Cédula de Identidad","Delegación","Cargo","Teléfono",
                      "Género","Sexo","Rango de Edad","Seleccionar"]],
            hide_index=True,
            use_container_width=True,
            column_config={
                "Nº": st.column_config.NumberColumn("Nº", disabled=True),
                "Seleccionar": st.column_config.CheckboxColumn("Seleccionar", help="Marca para eliminar"),
                "Género": st.column_config.SelectboxColumn("Género", options=["F","M","LGBTIQ+"]),
                "Sexo": st.column_config.SelectboxColumn("Sexo", options=["H","M","I"]),
                "Rango de Edad": st.column_config.SelectboxColumn("Rango de Edad",
                    options=["18 a 35 años","36 a 64 años","65 años o más"])
            },
            key="tabla_admin_editable"
        )

        c1, c2, c3, c4 = st.columns([1.2, 1.2, 1.2, 2])
        btn_save = c1.button("💾 Guardar cambios", use_container_width=True)
        btn_delete = c2.button("🗑️ Eliminar seleccionados", use_container_width=True)
        confirm_all = c4.checkbox("Confirmar vaciado total", value=False)
        btn_clear = c3.button("🧹 Vaciar todos", use_container_width=True)

        if btn_save:
            changes = 0
            for idx in edited.index:
                if idx >= len(df_view):
                    continue
                row_orig = df_view.loc[idx]
                row_new  = edited.loc[idx]
                fields = ["Nombre","Cédula de Identidad","Delegación","Cargo","Teléfono","Género","Sexo","Rango de Edad"]
                if any(str(row_orig[f]) != str(row_new[f]) for f in fields):
                    update_row_by_rownum(int(row_orig["rownum"]), {f: row_new[f] for f in fields})
                    changes += 1
            st.success(f"Se guardaron {changes} cambio(s).") if changes else st.info("No hay cambios para guardar.")
            if changes:
                st.rerun()

        if btn_delete:
            idx_sel = edited.index[edited["Seleccionar"] == True].tolist()
            rownums = df_view.iloc[idx_sel]["rownum"].tolist()
            if rownums:
                delete_rows_by_rownums(rownums)
                st.success(f"Eliminadas {len(rownums)} fila(s).")
                st.rerun()
            else:
                st.info("No hay filas seleccionadas para eliminar.")

        if btn_clear:
            if confirm_all:
                delete_all_rows()
                st.success("Se vaciaron todos los registros.")
                st.rerun()
            else:
                st.warning("Marca 'Confirmar vaciado total' para continuar.")

    # ===== Descarga Excel =====
    st.markdown("### ⬇️ Descarga")
    # Conjunto para exportar: SIEMPRE el filtrado actual
    df_for_export = df_view.drop(columns=["Nº","rownum"]) if not df_view.empty else df_view

    # ---------- EXCEL ----------
    def build_excel_oficial_single(
        fecha: date, lugar: str, hora_ini: time, hora_fin: time,
        estrategia: str, delegacion_hdr: str, rows_df: pd.DataFrame,
        anotaciones_txt: str, acuerdos_txt: str, firmante: str
    ) -> bytes:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
        except Exception:
            st.error("Falta 'openpyxl' en requirements.txt")
            return b""

        azul_banda = "1F3B73"
        gris_head  = "D9D9D9"
        celda_fill = PatternFill("solid", fgColor=gris_head)
        th_font    = Font(bold=True)
        title_font = Font(bold=True, size=12)
        h1_font    = Font(bold=True, size=14)
        center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
        right      = Alignment(horizontal="right",  vertical="center")
        left       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        thin       = Side(style="thin", color="000000")
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

        def outline_box(r1, c1, r2, c2):
            for c in range(c1, c2+1):
                t = ws.cell(row=r1, column=c)
                t.border = Border(top=thin, left=t.border.left, right=t.border.right, bottom=t.border.bottom)
                b = ws.cell(row=r2, column=c)
                b.border = Border(bottom=thin, left=b.border.left, right=b.border.right, top=b.border.top)
            for r in range(r1, r2+1):
                l = ws.cell(row=r, column=c1)
                l.border = Border(left=thin, top=l.border.top, right=l.border.right, bottom=l.border.bottom)
                rgt = ws.cell(row=r, column=c2)
                rgt.border = Border(right=thin, top=rgt.border.top, left=rgt.border.left, bottom=rgt.border.bottom)

        def box_all(r1, c1, r2, c2):
            for r in range(r1, r2+1):
                for c in range(c1, c2+1):
                    ws.cell(row=r, column=c).border = border_all

        MESES_ES = ["enero","febrero","marzo","abril","mayo","junio","julio","agosto",
                    "septiembre","octubre","noviembre","diciembre"]
        mes_es = MESES_ES[fecha.month-1]

        wb = Workbook()
        ws = wb.active; ws.title = "Lista"
        ws.sheet_view.showGridLines = False

        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = ws.page_margins.right = 0.3
        ws.page_margins.top = ws.page_margins.bottom = 0.4

        widths = {"A": 2, "B": 6, "C": 22, "D": 22, "E": 22, "F": 18, "G": 22,
                  "H": 20, "I": 16, "J": 6, "K": 6, "L": 10, "M": 6, "N": 6, "O": 6,
                  "P": 14, "Q": 14, "R": 14, "S": 16}
        for col, w in widths.items(): ws.column_dimensions[col].width = w

        ws.row_dimensions[1].height = 8
        ws.row_dimensions[3].height = 50
        ws.row_dimensions[4].height = 22
        ws.row_dimensions[5].height = 18
        ws.row_dimensions[6].height = 14

        # Títulos
        ws.merge_cells("B3:S3"); ws["B3"].value = "Modelo de Gestión Policial de Fuerza Pública"; ws["B3"].alignment=center; ws["B3"].font=h1_font
        ws.merge_cells("B4:S4"); ws["B4"].value = "Lista de Asistencia & Minuta"; ws["B4"].alignment=center; ws["B4"].font=h1_font
        ws.merge_cells("B5:S5"); ws["B5"].value = "Consecutivo:"; ws["B5"].alignment=center; ws["B5"].font=title_font
        ws.merge_cells("B6:S6"); ws["B6"].fill = PatternFill("solid", fgColor=azul_banda)

        outline_box(1, 2, 6, 19)  # marco superior

        # Encabezado superior
        ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=4)
        ws.merge_cells(start_row=7, start_column=5, end_row=7, end_column=9)
        ws.merge_cells(start_row=7, start_column=10, end_row=7, end_column=15)
        ws.merge_cells(start_row=7, start_column=16, end_row=7, end_column=19)
        ws["B7"].value = f"Fecha: {fecha.day} {mes_es} {fecha.year}"; ws["B7"].font = title_font; ws["B7"].alignment = left
        ws["E7"].value = f"Lugar:  {lugar}" if lugar else "Lugar: "; ws["E7"].font = title_font; ws["E7"].alignment = left
        ws["J7"].value = f"Hora Inicio: {hora_ini.strftime('%H:%M')}"; ws["J7"].alignment = center
        ws["P7"].value = f"Hora Finalización: {hora_fin.strftime('%H:%M')}"; ws["P7"].alignment = center
        box_all(7, 2, 7, 4); box_all(7, 5, 7, 9); box_all(7, 10, 7, 15); box_all(7, 16, 7, 19)

        # Estrategia
        ws.merge_cells(start_row=8, start_column=2, end_row=8, end_column=3)
        ws.merge_cells(start_row=8, start_column=4, end_row=8, end_column=9)
        ws["B8"].value = "Estrategia o Programa:"; ws["B8"].alignment = left
        ws["D8"].value = estrategia; ws["D8"].alignment = left
        box_all(8, 2, 8, 3); box_all(8, 4, 8, 9)

        # Actividad
        ws.merge_cells(start_row=8, start_column=10, end_row=9, end_column=19)
        ws["J8"].value = "ACTIVIDAD: Reunión Virtual de Seguimiento de líneas de acción, acciones estratégicas, indicadores y metas."
        ws["J8"].alignment = left
        outline_box(8, 10, 9, 19)

        # Delegación Policial (IZQUIERDA)
        ws.merge_cells(start_row=9, start_column=2, end_row=9, end_column=3)
        ws.merge_cells(start_row=9, start_column=4, end_row=9, end_column=9)
        ws["B9"].value = "Dirección / Delegación Policial:"; ws["B9"].alignment = left
        ws["D9"].value = delegacion_hdr; ws["D9"].alignment = left
        box_all(9, 2, 9, 3); box_all(9, 4, 9, 9)

        # Encabezado de la tabla
        ws["B10"].value = ""
        ws.merge_cells("C10:E11"); ws["C10"].value = "Nombre"

        ws["F10"].value = "Cédula de Identidad"
        ws["G10"].value = "Delegación"
        ws["H10"].value = "Cargo"
        ws["I10"].value = "Teléfono"
        ws.merge_cells("J10:L10"); ws["J10"].value = "Género"
        ws.merge_cells("M10:O10"); ws["M10"].value = "Sexo (Hombre, Mujer o Intersex)"
        ws.merge_cells("P10:R10"); ws["P10"].value = "Rango de Edad"
        ws["S10"].value = "FIRMA"

        for rng in ["C10:E11","J10:L10","M10:O10","P10:R10"]:
            c = ws[rng.split(":")[0]]; c.font = Font(bold=True); c.alignment = center; c.fill = celda_fill
        for cell in ["F10","G10","H10","I10","S10"]:
            ws[cell].font = Font(bold=True); ws[cell].alignment = center; ws[cell].fill = celda_fill

        ws["J11"], ws["K11"], ws["L11"] = "F", "M", "LGBTIQ+"
        ws["M11"], ws["N11"], ws["O11"] = "H", "M", "I"
        ws["P11"], ws["Q11"], ws["R11"] = "18 a 35 años", "36 a 64 años", "65 años o más"
        for cell in ["J11","K11","L11","M11","N11","O11","P11","Q11","R11"]:
            ws[cell].font = Font(bold=True); ws[cell].alignment = center; ws[cell].fill = celda_fill

        for r in range(10, 12):
            for c in range(2, 20):
                ws.cell(row=r, column=c).border = border_all

        ws.freeze_panes = "A12"

        # Filas
        start_row = 12
        for i, (_, row) in enumerate(rows_df.iterrows()):
            r = start_row + i
            ws[f"B{r}"].value = i + 1
            ws[f"B{r}"].alignment = right

            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
            ws[f"C{r}"].value = str(row.get("Nombre",""))
            ws[f"C{r}"].alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")

            ws[f"F{r}"].value = str(row.get("Cédula de Identidad",""))
            ws[f"G{r}"].value = str(row.get("Delegación",""))
            ws[f"H{r}"].value = str(row.get("Cargo",""))
            ws[f"I{r}"].value = str(row.get("Teléfono",""))

            for col in ["J","K","L","M","N","O","P","Q","R"]:
                ws[f"{col}{r}"].value = ""

            g = (row.get("Género","") or "").strip()
            if g == "F": ws[f"J{r}"].value = "X"
            elif g == "M": ws[f"K{r}"].value = "X"
            elif g == "LGBTIQ+": ws[f"L{r}"].value = "X"

            s = (row.get("Sexo","") or "").strip()
            if s == "H": ws[f"M{r}"].value = "X"
            elif s == "M": ws[f"N{r}"].value = "X"
            elif s == "I": ws[f"O{r}"].value = "X"

            e = (row.get("Rango de Edad","") or "").strip()
            if e.startswith("18"): ws[f"P{r}"].value = "X"
            elif e.startswith("36"): ws[f"Q{r}"].value = "X"
            elif e.startswith("65"): ws[f"R{r}"].value = "X"

            ws[f"S{r}"].value = "Virtual"

            for c in range(2, 20):
                ws.cell(row=r, column=c).border = border_all

        # Anotaciones / Acuerdos
        last_data_row = start_row + len(rows_df) - 1 if len(rows_df) > 0 else 11
        notes_top = max(25, last_data_row + 2)
        notes_height = 14

        ws.merge_cells(start_row=notes_top, start_column=2, end_row=notes_top, end_column=10)
        ws.merge_cells(start_row=notes_top, start_column=12, end_row=notes_top, end_column=19)
        ws[f"B{notes_top}"].value = "Anotaciones Generales."; ws[f"B{notes_top}"].alignment = center
        ws[f"L{notes_top}"].value = "Acuerdos."; ws[f"L{notes_top}"].alignment = center
        ws[f"B{notes_top}"].font = Font(bold=True); ws[f"L{notes_top}"].font = Font(bold=True)
        ws[f"B{notes_top}"].fill = celda_fill; ws[f"L{notes_top}"].fill = celda_fill

        outline_box(notes_top+1, 2, notes_top+notes_height, 10)
        outline_box(notes_top+1, 12, notes_top+notes_height, 19)

        ws.merge_cells(start_row=notes_top+1, start_column=2, end_row=notes_top+notes_height, end_column=10)
        ws[f"B{notes_top+1}"].alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        if anotaciones_txt.strip(): ws[f"B{notes_top+1}"].value = anotaciones_txt.strip()

        ws.merge_cells(start_row=notes_top+1, start_column=12, end_row=notes_top+notes_height, end_column=19)
        ws[f"L{notes_top+1}"].alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        if acuerdos_txt.strip(): ws[f"L{notes_top+1}"].value = acuerdos_txt.strip()

        # Pie con firma: nombre SOBRE la línea y etiqueta "Nombre" debajo
        row_pie = notes_top + notes_height + 2
        ws.merge_cells(start_row=row_pie, start_column=2, end_row=row_pie, end_column=10)
        ws[f"B{row_pie}"].value = f"Se Finaliza la Reunión a:   {hora_fin.strftime('%H:%M')}"
        ws[f"B{row_pie}"].alignment = left

        row_firma = row_pie + 3
        thin_line = Side(style="thin", color="000000")
        sig_c1, sig_c2 = 4, 10  # D..J

        # Línea de firma
        ws.merge_cells(start_row=row_firma, start_column=sig_c1, end_row=row_firma, end_column=sig_c2)
        for c in range(sig_c1, sig_c2 + 1):
            ws.cell(row=row_firma, column=c).border = Border(bottom=thin_line)

        from openpyxl.utils import get_column_letter
        col = get_column_letter(sig_c1)

        # Altura para que el texto “toque” la línea
        ws.row_dimensions[row_firma].height = 24

        # Nombre SOBRE la línea (misma fila de la línea, alineado abajo)
        texto_firma = firmante.strip() if (firmante and firmante.strip()) else ""
        ws[f"{col}{row_firma}"].value = texto_firma
        ws[f"{col}{row_firma}"].alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=False)

        # Etiqueta DEBAJO de la línea
        ws.merge_cells(start_row=row_firma+1, start_column=sig_c1, end_row=row_firma+1, end_column=sig_c2)
        ws[f"{col}{row_firma+1}"].value = "Nombre"
        ws[f"{col}{row_firma+1}"].alignment = Alignment(horizontal="center", wrap_text=False)

        ws.protection.sheet = True
        ws.protection.selectLockedCells = True
        ws.protection.selectUnlockedCells = True

        bio = BytesIO(); wb.save(bio); return bio.getvalue()

    # Botón de descarga Excel
    if st.button("📥 Generar Excel oficial", use_container_width=True, type="primary"):
        xls_bytes = build_excel_oficial_single(
            fecha_evento, lugar, hora_inicio, hora_fin, estrategia, delegacion_hdr,
            df_for_export, anotaciones, acuerdos, firmante_nombre
        )
        if xls_bytes:
            st.download_button(
                "⬇️ Descargar Excel (una sola hoja)",
                data=xls_bytes,
                file_name=f"Lista_Asistencia_Oficial_{date.today():%Y%m%d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )






