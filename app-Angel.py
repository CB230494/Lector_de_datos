# =========================
# 📋 Asistencia – Público + Admin
# Fecha/hora dispositivo + respaldo servidor
# =========================

import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import date, time, datetime
from typing import List
import uuid

import gspread
from streamlit_javascript import st_javascript

try:
    from zoneinfo import ZoneInfo
except Exception:
    ZoneInfo = None

st.set_page_config(page_title="Asistencia – Registro y Admin", layout="wide")

# ---------- Google Sheets ----------

SHEET_ID = "1vzGRJrlUzaCdhJAQBa6i94RE2QxnKqFvXpch9HF4TO8"
SHEET_NAME = "Hoja 1"

HEADER = [
    "id_registro",
    "nombre",
    "cedula",
    "delegacion",
    "cargo",
    "telefono",
    "genero",
    "sexo",
    "edad",
    "fecha_dispositivo",
    "hora_dispositivo",
    "timestamp_dispositivo",
    "zona_horaria_dispositivo",
    "fecha_servidor",
    "hora_servidor",
    "timestamp_servidor"
]

DELEGACIONES = [
    "Alajuela Sur", "Alajuela Norte", "San Ramón", "Grecia", "San Mateo",
    "Atenas", "Naranjo", "Palmares", "Poas", "Orotina", "Sarchí",
]

def get_now_cr():
    if ZoneInfo:
        return datetime.now(ZoneInfo("America/Costa_Rica"))
    return datetime.now()

def get_device_info():
    data = st_javascript("""
    (() => {
        const now = new Date();
        return {
            fecha: now.toLocaleDateString('es-CR'),
            hora: now.toLocaleTimeString('es-CR'),
            timestamp: now.toISOString(),
            zona: Intl.DateTimeFormat().resolvedOptions().timeZone
        };
    })()
    """)
    return data if isinstance(data, dict) else {}

def get_safe_device_info(device_info):
    if device_info:
        return device_info

    server_now = get_now_cr()
    return {
        "fecha": server_now.strftime("%d/%m/%Y"),
        "hora": server_now.strftime("%H:%M:%S"),
        "timestamp": server_now.isoformat(),
        "zona": "No detectada - respaldo servidor"
    }

def _sa_key():
    try:
        sa = st.secrets["gcp_service_account"]
        return sa.get("client_email", "") + "|" + sa.get("project_id", "")
    except Exception:
        return ""

@st.cache_resource(show_spinner=False)
def _get_ws_cached(sheet_id: str, sheet_name: str, sa_key: str):
    if "gcp_service_account" not in st.secrets:
        raise RuntimeError("Falta el bloque [gcp_service_account] en .streamlit/secrets.toml")

    gc = gspread.service_account_from_dict(st.secrets["gcp_service_account"])
    sh = gc.open_by_key(sheet_id)

    try:
        ws = sh.worksheet(sheet_name)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=sheet_name, rows=2000, cols=len(HEADER))
        ws.update("A1:P1", [HEADER])
        try:
            ws.freeze(rows=1)
        except Exception:
            pass

    first_row = [h.strip().lower() for h in ws.row_values(1)]

    if first_row != HEADER:
        ws.clear()
        ws.update("A1:P1", [HEADER])
        try:
            ws.freeze(rows=1)
        except Exception:
            pass

    return ws

def _get_ws():
    return _get_ws_cached(SHEET_ID, SHEET_NAME, _sa_key())

def init_db():
    _get_ws()

# ---------- CRUD ----------

def insert_row(row: dict):
    ws = _get_ws()

    telefono = row.get("Teléfono", "")
    if telefono and not str(telefono).startswith("'"):
        telefono = "'" + str(telefono)

    device = row.get("_device", {}) or {}
    server_now = get_now_cr()

    payload = [
        str(uuid.uuid4()),
        row.get("Nombre", ""),
        row.get("Cédula de Identidad", ""),
        row.get("Delegación", ""),
        row.get("Cargo", ""),
        telefono,
        row.get("Género", ""),
        row.get("Sexo", ""),
        row.get("Rango de Edad", ""),
        device.get("fecha", ""),
        device.get("hora", ""),
        device.get("timestamp", ""),
        device.get("zona", ""),
        server_now.strftime("%d/%m/%Y"),
        server_now.strftime("%H:%M:%S"),
        server_now.strftime("%Y-%m-%d %H:%M:%S"),
    ]

    ws.append_row(payload, value_input_option="USER_ENTERED")

def fetch_all_df(include_rownum=True) -> pd.DataFrame:
    ws = _get_ws()
    values = ws.get_all_values()

    base_cols = [
        "Nº", "Nombre", "Cédula de Identidad", "Delegación", "Cargo", "Teléfono",
        "Género", "Sexo", "Rango de Edad", "Fecha Dispositivo", "Hora Dispositivo",
        "Timestamp Dispositivo", "Zona Horaria Dispositivo",
        "Fecha Servidor", "Hora Servidor", "Timestamp Servidor"
    ]

    if len(values) < 2:
        cols = base_cols.copy()
        if include_rownum:
            cols.insert(1, "rownum")
        return pd.DataFrame(columns=cols)

    header = [h.strip().lower() for h in values[0]]
    data_rows = values[1:]

    name_map = {
        "id_registro": "ID Registro",
        "nombre": "Nombre",
        "cedula": "Cédula de Identidad",
        "delegacion": "Delegación",
        "cargo": "Cargo",
        "telefono": "Teléfono",
        "genero": "Género",
        "sexo": "Sexo",
        "edad": "Rango de Edad",
        "fecha_dispositivo": "Fecha Dispositivo",
        "hora_dispositivo": "Hora Dispositivo",
        "timestamp_dispositivo": "Timestamp Dispositivo",
        "zona_horaria_dispositivo": "Zona Horaria Dispositivo",
        "fecha_servidor": "Fecha Servidor",
        "hora_servidor": "Hora Servidor",
        "timestamp_servidor": "Timestamp Servidor",
    }

    records = []
    for idx, row in enumerate(data_rows, start=2):
        rec = {}
        for j, key in enumerate(header):
            if key in name_map:
                rec[name_map[key]] = row[j] if j < len(row) else ""
        rec["rownum"] = idx
        records.append(rec)

    df = pd.DataFrame(records)

    if df.empty:
        cols = base_cols.copy()
        if include_rownum:
            cols.insert(1, "rownum")
        return pd.DataFrame(columns=cols)

    cols_order = [
        "rownum", "ID Registro", "Nombre", "Cédula de Identidad", "Delegación",
        "Cargo", "Teléfono", "Género", "Sexo", "Rango de Edad",
        "Fecha Dispositivo", "Hora Dispositivo", "Timestamp Dispositivo",
        "Zona Horaria Dispositivo", "Fecha Servidor", "Hora Servidor",
        "Timestamp Servidor"
    ]

    df = df[[c for c in cols_order if c in df.columns]]
    df.insert(0, "Nº", range(1, len(df) + 1))

    if not include_rownum and "rownum" in df.columns:
        df = df.drop(columns=["rownum"])

    return df

def update_row_by_rownum(rownum: int, row: dict):
    ws = _get_ws()

    current_values = ws.row_values(rownum)
    while len(current_values) < len(HEADER):
        current_values.append("")

    payload = [
        current_values[0],
        row.get("Nombre", ""),
        row.get("Cédula de Identidad", ""),
        row.get("Delegación", ""),
        row.get("Cargo", ""),
        row.get("Teléfono", ""),
        row.get("Género", ""),
        row.get("Sexo", ""),
        row.get("Rango de Edad", ""),
        current_values[9],
        current_values[10],
        current_values[11],
        current_values[12],
        current_values[13],
        current_values[14],
        current_values[15],
    ]

    ws.update(f"A{rownum}:P{rownum}", [payload], value_input_option="USER_ENTERED")

def delete_rows_by_rownums(rownums: List[int]):
    if not rownums:
        return
    ws = _get_ws()
    for r in sorted(rownums, reverse=True):
        ws.delete_rows(r)

def delete_all_rows():
    ws = _get_ws()
    used_rows = len(ws.get_all_values())
    if used_rows >= 2:
        ws.batch_clear([f"A2:P{used_rows}"])

# ---------- Inicializar ----------

try:
    init_db()
except Exception as e:
    st.error("Error conectando a Google Sheets. Verifica permisos, secrets y nombre de hoja.")
    st.exception(e)
    st.stop()

# ---------- Login admin ----------

if "is_admin" not in st.session_state:
    st.session_state.is_admin = False

with st.sidebar:
    st.markdown("### 🔐 Acceso administrador")

    if not st.session_state.is_admin:
        pwd = st.text_input("Contraseña", type="password", placeholder="••••••••")
        if st.button("Ingresar"):
            if pwd == "Region223":
                st.session_state.is_admin = True
                st.success("Acceso concedido.")
                st.rerun()
            else:
                st.error("Contraseña incorrecta.")
    else:
        st.success("Sesión de administrador activa")
        if st.button("Cerrar sesión"):
            st.session_state.is_admin = False
            st.rerun()

# ---------- Público ----------

st.markdown("# 📋 Asistencia – Registro")

device_info = get_device_info()

if device_info:
    st.info(
        f"Fecha y hora del dispositivo: "
        f"{device_info.get('fecha', '')} {device_info.get('hora', '')} "
        f"({device_info.get('zona', '')})"
    )
else:
    st.info(
        "Fecha y hora del dispositivo en proceso de detección. "
        "Si no carga, el sistema usará la hora del servidor como respaldo."
    )

st.markdown("### ➕ Agregar")

with st.form("form_asistencia_publico", clear_on_submit=True):
    c1, c2, c3 = st.columns([1.2, 1, 1])

    nombre = c1.text_input("Nombre")
    cedula = c2.text_input("Cédula de Identidad")

    opciones_deleg = ["— Selecciona una delegación —"] + DELEGACIONES
    sel_deleg = c3.selectbox("Delegación", opciones_deleg, index=0)
    delegacion_sel = "" if sel_deleg == opciones_deleg[0] else sel_deleg

    c4, c5 = st.columns([1, 1])
    cargo = c4.text_input("Cargo")
    telefono = c5.text_input("Teléfono")

    gcol, scol, ecol = st.columns([1.1, 1.5, 1.5])

    genero = gcol.radio("Género", ["F", "M", "LGBTIQ+"], horizontal=True)
    sexo = scol.radio("Sexo (Hombre, Mujer o Intersex)", ["H", "M", "I"], horizontal=True)
    edad = ecol.radio(
        "Rango de Edad",
        ["18 a 35 años", "36 a 64 años", "65 años o más"],
        horizontal=True
    )

    submitted = st.form_submit_button("➕ Agregar", use_container_width=True)

    if submitted:
        if not nombre.strip():
            st.warning("Ingresa al menos el nombre.")
        else:
            device_final = get_safe_device_info(device_info)

            fila = {
                "Nombre": nombre.strip(),
                "Cédula de Identidad": cedula.strip(),
                "Delegación": delegacion_sel.strip(),
                "Cargo": cargo.strip(),
                "Teléfono": telefono.strip(),
                "Género": genero,
                "Sexo": sexo,
                "Rango de Edad": edad,
                "_device": device_final
            }

            insert_row(fila)

            if device_info:
                st.success("Registro guardado con fecha y hora del dispositivo.")
            else:
                st.success("Registro guardado con fecha y hora del servidor como respaldo.")

st.markdown("### 📥 Registros recibidos")

df_pub = fetch_all_df(include_rownum=False)

cols_publicas = [
    "Nº", "Nombre", "Cédula de Identidad", "Delegación", "Cargo", "Teléfono",
    "Género", "Sexo", "Rango de Edad", "Fecha Dispositivo", "Hora Dispositivo"
]

if not df_pub.empty:
    st.dataframe(
        df_pub[[c for c in cols_publicas if c in df_pub.columns]],
        use_container_width=True,
        hide_index=True
    )
else:
    st.info("Aún no hay registros guardados.")

# ---------- Admin ----------

if st.session_state.is_admin:
    st.markdown("---")
    st.markdown("# 🛠️ Panel del Administrador")

    df_all = fetch_all_df(include_rownum=True)

    if df_all.empty:
        st.info("Aún no hay registros guardados.")
        st.stop()

    delegs_existentes = sorted(
        [d for d in df_all["Delegación"].dropna().unique() if str(d).strip()],
        key=str.casefold
    )

    sel_filtros = st.multiselect(
        "Filtrar por Delegación",
        options=delegs_existentes,
        default=[],
        help="Vacío = todas. Puedes elegir varias delegaciones."
    )

    if not sel_filtros:
        df_view = df_all.copy().reset_index(drop=True)
    else:
        df_view = df_all[df_all["Delegación"].isin(sel_filtros)].reset_index(drop=True)

    st.markdown("### 🧾 Datos de encabezado para el Excel")

    col1, col2 = st.columns([1, 1])

    with col1:
        fecha_evento = st.date_input("Fecha", value=date.today())
        lugar = st.text_input("Lugar", value="")
        estrategia = st.text_input("Estrategia o Programa", value="Estrategia Sembremos Seguridad")

    with col2:
        hora_inicio = st.time_input("Hora Inicio", value=time(9, 0))
        hora_fin = st.time_input("Hora Finalización", value=time(12, 10))
        delegacion_hdr = st.text_input("Dirección / Delegación Policial", value="")
        firmante_nombre = st.text_input("Nombre de quien firma (opcional)", value="")

    st.markdown("### 📝 Anotaciones y Acuerdos para el Excel")

    a_col, b_col = st.columns(2)

    anotaciones = a_col.text_area(
        "Anotaciones Generales",
        height=220,
        placeholder="Escribe las anotaciones generales…"
    )

    acuerdos = b_col.text_area(
        "Acuerdos",
        height=220,
        placeholder="Escribe los acuerdos…"
    )

    st.markdown("### 👥 Registros y edición")

    if df_view.empty:
        st.info("No hay registros para el filtro seleccionado.")
    else:
        editable = df_view.copy()
        editable["Seleccionar"] = False

        edited = st.data_editor(
            editable[
                [
                    "Nº", "Nombre", "Cédula de Identidad", "Delegación", "Cargo",
                    "Teléfono", "Género", "Sexo", "Rango de Edad",
                    "Fecha Dispositivo", "Hora Dispositivo",
                    "Fecha Servidor", "Hora Servidor",
                    "Seleccionar"
                ]
            ],
            hide_index=True,
            use_container_width=True,
            column_config={
                "Nº": st.column_config.NumberColumn("Nº", disabled=True),
                "Fecha Dispositivo": st.column_config.TextColumn("Fecha Dispositivo", disabled=True),
                "Hora Dispositivo": st.column_config.TextColumn("Hora Dispositivo", disabled=True),
                "Fecha Servidor": st.column_config.TextColumn("Fecha Servidor", disabled=True),
                "Hora Servidor": st.column_config.TextColumn("Hora Servidor", disabled=True),
                "Seleccionar": st.column_config.CheckboxColumn("Seleccionar", help="Marca para eliminar"),
                "Género": st.column_config.SelectboxColumn("Género", options=["F", "M", "LGBTIQ+"]),
                "Sexo": st.column_config.SelectboxColumn("Sexo", options=["H", "M", "I"]),
                "Rango de Edad": st.column_config.SelectboxColumn(
                    "Rango de Edad",
                    options=["18 a 35 años", "36 a 64 años", "65 años o más"]
                )
            },
            key="tabla_admin_editable"
        )

        c1, c2, c3, c4 = st.columns([1.2, 1.2, 1.2, 2])

        btn_save = c1.button("💾 Guardar cambios", use_container_width=True)
        btn_delete = c2.button("🗑️ Eliminar seleccionados", use_container_width=True)
        btn_clear = c3.button("🧹 Vaciar todos", use_container_width=True)
        confirm_all = c4.checkbox("Confirmar vaciado total", value=False)

        if btn_save:
            changes = 0

            for idx in edited.index:
                if idx >= len(df_view):
                    continue

                row_orig = df_view.loc[idx]
                row_new = edited.loc[idx]

                fields = [
                    "Nombre", "Cédula de Identidad", "Delegación", "Cargo",
                    "Teléfono", "Género", "Sexo", "Rango de Edad"
                ]

                if any(str(row_orig[f]) != str(row_new[f]) for f in fields):
                    update_row_by_rownum(
                        int(row_orig["rownum"]),
                        {f: row_new[f] for f in fields}
                    )
                    changes += 1

            if changes:
                st.success(f"Se guardaron {changes} cambio(s).")
                st.rerun()
            else:
                st.info("No hay cambios para guardar.")

        if btn_delete:
            idx_sel = edited.index[edited["Seleccionar"] == True].tolist()
            rownums = df_view.iloc[idx_sel]["rownum"].tolist()

            if rownums:
                delete_rows_by_rownums(rownums)
                st.success(f"Eliminadas {len(rownums)} fila(s).")
                st.rerun()
            else:
                st.info("No hay filas seleccionadas para eliminar.")

        if btn_clear:
            if confirm_all:
                delete_all_rows()
                st.success("Se vaciaron todos los registros.")
                st.rerun()
            else:
                st.warning("Marca 'Confirmar vaciado total' para continuar.")

    # ---------- Excel oficial ----------

    st.markdown("### ⬇️ Descarga")

    export_cols = [
        "Nombre", "Cédula de Identidad", "Delegación", "Cargo", "Teléfono",
        "Género", "Sexo", "Rango de Edad",
        "Fecha Dispositivo", "Hora Dispositivo",
        "Fecha Servidor", "Hora Servidor"
    ]

    df_for_export = df_view[[c for c in export_cols if c in df_view.columns]].copy()

    def build_excel_oficial_single(
        fecha: date,
        lugar: str,
        hora_ini: time,
        hora_fin: time,
        estrategia: str,
        delegacion_hdr: str,
        rows_df: pd.DataFrame,
        anotaciones_txt: str,
        acuerdos_txt: str,
        firmante: str
    ) -> bytes:

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
            from openpyxl.drawing.image import Image as XLImage
            from pathlib import Path as _Path
        except Exception:
            st.error("Falta 'openpyxl' y/o 'Pillow' en requirements.txt")
            return b""

        gris_head = "D9D9D9"
        celda_fill = PatternFill("solid", fgColor=gris_head)
        th_font = Font(bold=True)
        title_font = Font(bold=True, size=12)
        h1_font = Font(bold=True, size=14)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        right = Alignment(horizontal="right", vertical="center")
        left = Alignment(horizontal="left", vertical="top", wrap_text=True)
        thin = Side(style="thin", color="000000")
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

        def outline_box(r1, c1, r2, c2):
            for c in range(c1, c2 + 1):
                t = ws.cell(row=r1, column=c)
                t.border = Border(
                    top=thin,
                    left=t.border.left,
                    right=t.border.right,
                    bottom=t.border.bottom
                )

                b = ws.cell(row=r2, column=c)
                b.border = Border(
                    bottom=thin,
                    left=b.border.left,
                    right=b.border.right,
                    top=b.border.top
                )

            for r in range(r1, r2 + 1):
                l = ws.cell(row=r, column=c1)
                l.border = Border(
                    left=thin,
                    top=l.border.top,
                    right=l.border.right,
                    bottom=l.border.bottom
                )

                rgt = ws.cell(row=r, column=c2)
                rgt.border = Border(
                    right=thin,
                    top=rgt.border.top,
                    left=rgt.border.left,
                    bottom=rgt.border.bottom
                )

        def box_all(r1, c1, r2, c2):
            for r in range(r1, r2 + 1):
                for c in range(c1, c2 + 1):
                    ws.cell(row=r, column=c).border = border_all

        MESES_ES = [
            "enero", "febrero", "marzo", "abril", "mayo", "junio",
            "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"
        ]

        mes_es = MESES_ES[fecha.month - 1]

        wb = Workbook()
        ws = wb.active
        ws.title = "Lista"
        ws.sheet_view.showGridLines = False

        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = ws.page_margins.right = 0.3
        ws.page_margins.top = ws.page_margins.bottom = 0.4

        widths = {
            "A": 2, "B": 6, "C": 26, "D": 22, "E": 22,
            "F": 18, "G": 24, "H": 28, "I": 20,
            "J": 6, "K": 6, "L": 10, "M": 6, "N": 6, "O": 6,
            "P": 14, "Q": 14, "R": 14, "S": 16
        }

        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        ws.row_dimensions[1].height = 8
        ws.row_dimensions[3].height = 50
        ws.row_dimensions[4].height = 22
        ws.row_dimensions[5].height = 18
        ws.row_dimensions[6].height = 14

        try:
            if _Path("logo_izq.png").exists():
                img = XLImage("logo_izq.png")
                target_h = 72
                ratio = target_h / img.height
                img.height = target_h
                img.width = int(img.width * ratio)
                ws.add_image(img, "D3")

            if _Path("logo_der.png").exists():
                img2 = XLImage("logo_der.png")
                target_h2 = 72
                ratio2 = target_h2 / img2.height
                img2.height = target_h2
                img2.width = int(img2.width * ratio2)
                ws.add_image(img2, "O3")
        except Exception:
            pass

        ws.merge_cells("B3:S3")
        ws["B3"].value = "Modelo de Gestión Policial de Fuerza Pública"
        ws["B3"].alignment = center
        ws["B3"].font = h1_font

        ws.merge_cells("B4:S4")
        ws["B4"].value = "Lista de Asistencia & Minuta"
        ws["B4"].alignment = center
        ws["B4"].font = h1_font

        ws.merge_cells("B5:S5")
        ws["B5"].value = "Consecutivo:"
        ws["B5"].alignment = center
        ws["B5"].font = title_font

        ws.merge_cells("B6:S6")
        ws["B6"].fill = PatternFill("solid", fgColor="1F3B73")
        outline_box(1, 2, 6, 19)

        ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=4)
        ws.merge_cells(start_row=7, start_column=5, end_row=7, end_column=9)
        ws.merge_cells(start_row=7, start_column=10, end_row=7, end_column=15)
        ws.merge_cells(start_row=7, start_column=16, end_row=7, end_column=19)

        ws["B7"].value = f"Fecha: {fecha.day} {mes_es} {fecha.year}"
        ws["B7"].font = title_font
        ws["B7"].alignment = left

        ws["E7"].value = f"Lugar: {lugar}" if lugar else "Lugar:"
        ws["E7"].font = title_font
        ws["E7"].alignment = left

        ws["J7"].value = f"Hora Inicio: {hora_ini.strftime('%H:%M')}"
        ws["J7"].alignment = center

        ws["P7"].value = f"Hora Finalización: {hora_fin.strftime('%H:%M')}"
        ws["P7"].alignment = center

        box_all(7, 2, 7, 4)
        box_all(7, 5, 7, 9)
        box_all(7, 10, 7, 15)
        box_all(7, 16, 7, 19)

        ws.merge_cells(start_row=8, start_column=2, end_row=8, end_column=3)
        ws.merge_cells(start_row=8, start_column=4, end_row=8, end_column=9)

        ws["B8"].value = "Estrategia o Programa:"
        ws["B8"].alignment = left

        ws["D8"].value = estrategia
        ws["D8"].alignment = left

        box_all(8, 2, 8, 3)
        box_all(8, 4, 8, 9)

        ws.merge_cells(start_row=8, start_column=10, end_row=9, end_column=19)
        ws["J8"].value = (
            "ACTIVIDAD: Reunión Virtual de Seguimiento de líneas de acción, "
            "acciones estratégicas, indicadores y metas."
        )
        ws["J8"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        outline_box(8, 10, 9, 19)

        ws.merge_cells(start_row=9, start_column=2, end_row=9, end_column=3)
        ws.merge_cells(start_row=9, start_column=4, end_row=9, end_column=9)

        ws["B9"].value = "Dirección / Delegación Policial:"
        ws["B9"].alignment = left

        ws["D9"].value = delegacion_hdr
        ws["D9"].alignment = left

        box_all(9, 2, 9, 3)
        box_all(9, 4, 9, 9)

        ws["B10"].value = ""
        ws.merge_cells("C10:E11")
        ws["C10"].value = "Nombre"

        ws["F10"].value = "Cédula de Identidad"
        ws["G10"].value = "Delegación"
        ws["H10"].value = "Cargo"
        ws["I10"].value = "Teléfono"

        ws.merge_cells("J10:L10")
        ws["J10"].value = "Género"

        ws.merge_cells("M10:O10")
        ws["M10"].value = "Sexo (Hombre, Mujer o Intersex)"

        ws.merge_cells("P10:R10")
        ws["P10"].value = "Rango de Edad"

        ws["S10"].value = "FIRMA"

        for rng in ["C10:E11", "J10:L10", "M10:O10", "P10:R10"]:
            c = ws[rng.split(":")[0]]
            c.font = th_font
            c.alignment = center
            c.fill = celda_fill

        for cell in ["F10", "G10", "H10", "I10", "S10"]:
            ws[cell].font = th_font
            ws[cell].alignment = center
            ws[cell].fill = celda_fill

        ws["J11"], ws["K11"], ws["L11"] = "F", "M", "LGBTIQ+"
        ws["M11"], ws["N11"], ws["O11"] = "H", "M", "I"
        ws["P11"], ws["Q11"], ws["R11"] = "18 a 35 años", "36 a 64 años", "65 años o más"

        for cell in ["J11", "K11", "L11", "M11", "N11", "O11", "P11", "Q11", "R11"]:
            ws[cell].font = th_font
            ws[cell].alignment = center
            ws[cell].fill = celda_fill

        for r in range(10, 12):
            for c in range(2, 20):
                ws.cell(row=r, column=c).border = border_all

        ws.freeze_panes = "A12"

        start_row = 12

        for i, (_, row) in enumerate(rows_df.iterrows()):
            r = start_row + i

            ws[f"B{r}"].value = i + 1
            ws[f"B{r}"].alignment = right

            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
            ws[f"C{r}"].value = str(row.get("Nombre", ""))
            ws[f"C{r}"].alignment = left

            ws[f"F{r}"].value = str(row.get("Cédula de Identidad", ""))
            ws[f"G{r}"].value = str(row.get("Delegación", ""))
            ws[f"H{r}"].value = str(row.get("Cargo", ""))
            ws[f"I{r}"].value = str(row.get("Teléfono", ""))

            for col in ["F", "G", "H", "I"]:
                ws[f"{col}{r}"].alignment = left

            for col in ["J", "K", "L", "M", "N", "O", "P", "Q", "R"]:
                ws[f"{col}{r}"].value = ""

            g = (row.get("Género", "") or "").strip()
            if g == "F":
                ws[f"J{r}"].value = "X"
            elif g == "M":
                ws[f"K{r}"].value = "X"
            elif g == "LGBTIQ+":
                ws[f"L{r}"].value = "X"

            s = (row.get("Sexo", "") or "").strip()
            if s == "H":
                ws[f"M{r}"].value = "X"
            elif s == "M":
                ws[f"N{r}"].value = "X"
            elif s == "I":
                ws[f"O{r}"].value = "X"

            e = (row.get("Rango de Edad", "") or "").strip()
            if e.startswith("18"):
                ws[f"P{r}"].value = "X"
            elif e.startswith("36"):
                ws[f"Q{r}"].value = "X"
            elif e.startswith("65"):
                ws[f"R{r}"].value = "X"

            ws[f"S{r}"].value = "Virtual"

            for c in range(2, 20):
                ws.cell(row=r, column=c).border = border_all

        last_data_row = start_row + len(rows_df) - 1 if len(rows_df) > 0 else 11

        evidencia_top = last_data_row + 2

        ws.merge_cells(start_row=evidencia_top, start_column=2, end_row=evidencia_top, end_column=19)
        ws[f"B{evidencia_top}"].value = "Trazabilidad del registro electrónico de asistencia"
        ws[f"B{evidencia_top}"].font = th_font
        ws[f"B{evidencia_top}"].alignment = center
        ws[f"B{evidencia_top}"].fill = celda_fill
        box_all(evidencia_top, 2, evidencia_top, 19)

        evidencia_text_row = evidencia_top + 1
        ws.merge_cells(start_row=evidencia_text_row, start_column=2, end_row=evidencia_text_row + 2, end_column=19)
        ws[f"B{evidencia_text_row}"].value = (
            "Los registros de asistencia fueron capturados mediante formulario electrónico. "
            "El sistema almacena la fecha y hora reportada por el dispositivo utilizado para el registro. "
            "En caso de no poder detectar dicha información, se utiliza la fecha y hora del servidor como respaldo técnico. "
            "Adicionalmente, se conserva la fecha y hora del servidor como mecanismo complementario de trazabilidad y control."
        )
        ws[f"B{evidencia_text_row}"].alignment = left
        outline_box(evidencia_text_row, 2, evidencia_text_row + 2, 19)

        notes_top = evidencia_text_row + 5
        notes_height = 14

        ws.merge_cells(start_row=notes_top, start_column=2, end_row=notes_top, end_column=10)
        ws.merge_cells(start_row=notes_top, start_column=12, end_row=notes_top, end_column=19)

        ws[f"B{notes_top}"].value = "Anotaciones Generales."
        ws[f"L{notes_top}"].value = "Acuerdos."

        ws[f"B{notes_top}"].alignment = center
        ws[f"L{notes_top}"].alignment = center

        ws[f"B{notes_top}"].font = th_font
        ws[f"L{notes_top}"].font = th_font

        ws[f"B{notes_top}"].fill = celda_fill
        ws[f"L{notes_top}"].fill = celda_fill

        outline_box(notes_top + 1, 2, notes_top + notes_height, 10)
        outline_box(notes_top + 1, 12, notes_top + notes_height, 19)

        ws.merge_cells(start_row=notes_top + 1, start_column=2, end_row=notes_top + notes_height, end_column=10)
        ws[f"B{notes_top + 1}"].alignment = left

        if anotaciones_txt.strip():
            ws[f"B{notes_top + 1}"].value = anotaciones_txt.strip()

        ws.merge_cells(start_row=notes_top + 1, start_column=12, end_row=notes_top + notes_height, end_column=19)
        ws[f"L{notes_top + 1}"].alignment = left

        if acuerdos_txt.strip():
            ws[f"L{notes_top + 1}"].value = acuerdos_txt.strip()

        row_pie = notes_top + notes_height + 2

        ws.merge_cells(start_row=row_pie, start_column=2, end_row=row_pie, end_column=10)
        ws[f"B{row_pie}"].value = f"Se Finaliza la Reunión a:   {hora_fin.strftime('%H:%M')}"
        ws[f"B{row_pie}"].alignment = left

        row_firma = row_pie + 3
        thin_line = Side(style="thin", color="000000")
        sig_c1, sig_c2 = 4, 10

        ws.merge_cells(start_row=row_firma, start_column=sig_c1, end_row=row_firma, end_column=sig_c2)

        for c in range(sig_c1, sig_c2 + 1):
            ws.cell(row=row_firma, column=c).border = Border(bottom=thin_line)

        col = get_column_letter(sig_c1)
        ws.row_dimensions[row_firma].height = 24

        ws[f"{col}{row_firma}"].value = (firmante or "").strip()
        ws[f"{col}{row_firma}"].alignment = Alignment(horizontal="center", vertical="bottom")

        ws.merge_cells(start_row=row_firma + 1, start_column=sig_c1, end_row=row_firma + 1, end_column=sig_c2)
        ws[f"{col}{row_firma + 1}"].value = "Nombre"
        ws[f"{col}{row_firma + 1}"].alignment = Alignment(horizontal="center")

        ws.merge_cells(start_row=row_firma + 3, start_column=2, end_row=row_firma + 3, end_column=10)
        ws[f"B{row_firma + 3}"].value = "Cargo:"
        ws[f"B{row_firma + 3}"].alignment = left

        ws.merge_cells(start_row=row_firma + 5, start_column=12, end_row=row_firma + 5, end_column=19)
        ws[f"L{row_firma + 5}"].value = "Sello Policial"
        ws[f"L{row_firma + 5}"].alignment = Alignment(horizontal="right", vertical="center")

        ws.protection.sheet = True
        ws.protection.selectLockedCells = True
        ws.protection.selectUnlockedCells = True

        bio = BytesIO()
        wb.save(bio)
        return bio.getvalue()

    if st.button("📥 Generar Excel oficial", use_container_width=True, type="primary"):
        xls_bytes = build_excel_oficial_single(
            fecha_evento,
            lugar,
            hora_inicio,
            hora_fin,
            estrategia,
            delegacion_hdr,
            df_for_export,
            anotaciones,
            acuerdos,
            firmante_nombre
        )

        if xls_bytes:
            st.download_button(
                "⬇️ Descargar Excel oficial",
                data=xls_bytes,
                file_name=f"Lista_Asistencia_Oficial_{date.today():%Y%m%d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
